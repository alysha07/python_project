from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "menu.xlsx"


# ---------- ENCAPSULATION ----------

class MenuManagement:

    def __init__(self):
        self.__file_name = FILE_NAME
        self.create_file()

    # ---------- CREATE FILE ----------

    def add_item(self):
        wb, ws = self.load_sheet()

        item_name = input("Enter Item Name: ").strip()

        try:
            price = float(input("Enter Price: ").strip())
        except ValueError:
            print("Invalid Price!\n")
            return

        print("\nSelect Category:")
        print("1. Beverages")
        print("2. Starters")
        print("3. Main Course")
        print("4. Desserts")

        category_choice = input("Enter Category Choice: ")

        category_dict = {
            "1": "Beverages",
            "2": "Starters",
            "3": "Main Course",
            "4": "Desserts"
        }

        if category_choice not in category_dict:
            print("Invalid Category!\n")
            return

        category = category_dict[category_choice]

        # --- THE FIX IS HERE ---
        # Initialize with a default value
        food_type = "N/A" 

        # Only ask for Food Type IF it's Starters (2) or Main Course (3)
        if category_choice in ["2", "3"]:
            print("\nSelect Food Type:")
            print("1. Veg")
            print("2. Non-Veg")

            food_choice = input("Enter Food Type Choice: ")
            food_dict = {"1": "Veg", "2": "Non-Veg"}

            if food_choice not in food_dict:
                print("Invalid Food Type!\n")
                return
            food_type = food_dict[food_choice]
        
        # If it was 1 or 4, the code skips the 'if' block above 
        # and moves straight to appending the data.
        
        ws.append([item_name, price, category, food_type])
        
        try:
            wb.save(self.__file_name)
            print(f"Item '{item_name}' Added Successfully!\n")
        except PermissionError:
            print("❌ Error: Close the Excel file first!")

    # ---------- LOAD SHEET ----------

    
    def load_sheet(self):
        try:
            wb = load_workbook(self.__file_name)
            ws = wb.active
            return wb, ws
        except PermissionError:
            print("❌ Error: Permission Denied. Please close 'menu.xlsx' in Excel and try again.")
            exit() # Stop the script before it hits more errors



   
    # ---------- DISPLAY MENU ----------

    def display_menu(self):
        wb, ws = self.load_sheet()

        print("\n--- MENU ---")
        print(f"{'Item Name':<20}{'Price':<10}{'Category':<15}{'Food Type':<10}")
        print("-" * 60)

        for row in ws.iter_rows(min_row=2, values_only=True):
            item_name, price, category, food_type = row

            if item_name is not None:
                print(f"{str(item_name):<20}{str(price):<10}{str(category):<15}{str(food_type):<10}")

        print()

    # ---------- SEARCH ITEM ----------

    def search_item(self):
        wb, ws = self.load_sheet()

        search_name = input("Enter Item Name to Search: ")
        found = False

        for row in ws.iter_rows(min_row=2, values_only=True):
            item_name, price, category, food_type = row

            if item_name and item_name.strip().lower() == search_name.strip().lower():
                print(f"\nFound -> {item_name} | ₹{price} | {category} | {food_type}\n")
                found = True
                break

        if not found:
            print("Item Not Found!\n")

    # ---------- UPDATE ITEM ----------

    def update_item(self):
        wb, ws = self.load_sheet()

        search_name = input("Enter Item Name to Update: ")
        updated = False

        for row in range(2, ws.max_row + 1):
            item_name = ws.cell(row=row, column=1).value

            if item_name and item_name.strip().lower() == search_name.strip().lower():

                new_name = input("Enter New Item Name: ")

                try:
                    new_price = float(input("Enter New Price: "))
                except ValueError:
                    print("Invalid Price!\n")
                    return

                print("\nSelect New Category:")
                print("1. Beverages")
                print("2. Starters")
                print("3. Main Course")
                print("4. Desserts")

                category_choice = input("Enter Category Choice: ")

                category_dict = {
                    "1": "Beverages",
                    "2": "Starters",
                    "3": "Main Course",
                    "4": "Desserts"
                }

                if category_choice not in category_dict:
                    print("Invalid Category!\n")
                    return

                new_category = category_dict[category_choice]

                print("\nSelect Food Type:")
                print("1. Veg")
                print("2. Non-Veg")

                food_choice = input("Enter Food Type Choice: ")

                food_dict = {
                    "1": "Veg",
                    "2": "Non-Veg"
                }

                if food_choice not in food_dict:
                    print("Invalid Food Type!\n")
                    return

                new_food_type = food_dict[food_choice]

                ws.cell(row=row, column=1).value = new_name
                ws.cell(row=row, column=2).value = new_price
                ws.cell(row=row, column=3).value = new_category
                ws.cell(row=row, column=4).value = new_food_type

                updated = True
                break

        wb.save(self.__file_name)

        if updated:
            print("Item Updated Successfully!\n")
        else:
            print("Item Not Found!\n")

    # ---------- DELETE ITEM ----------

    def delete_item(self):
        wb, ws = self.load_sheet()

        search_name = input("Enter Item Name to Delete: ")
        found = False

        for row in range(2, ws.max_row + 1):
            item_name = ws.cell(row=row, column=1).value

            if item_name and item_name.strip().lower() == search_name.strip().lower():
                ws.delete_rows(row)
                found = True
                break

        wb.save(self.__file_name)

        if found:
            print("Item Deleted Successfully!\n")
        else:
            print("Item Not Found!\n")


