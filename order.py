import os
from openpyxl import Workbook, load_workbook
from reservation import TableReservation
from menu import MenuManagement
from billing import generate_bill
from notifications import review_request_email   # ← NEW

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


class OrderSystem:

    def __init__(self, reservation_obj, menu_obj):
        # Use SAME objects from main.py (IMPORTANT)
        self.reservation_obj = reservation_obj
        self.menu_obj = menu_obj
        self.order_file = os.path.join(BASE_DIR, "data", "orders.txt")
        self.order_xlsx = os.path.join(BASE_DIR, "data", "orders.xlsx")
        self.create_order_files()

    def create_order_files(self):
        data_dir = os.path.dirname(self.order_file)
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        if not os.path.exists(self.order_file):
            open(self.order_file, "w", encoding="utf-8").close()

        if not os.path.exists(self.order_xlsx):
            wb = Workbook()
            ws = wb.active
            ws.append(["Order ID", "Customer", "Table", "Items"])
            wb.save(self.order_xlsx)

    # ---------- GENERATE ORDER ID ----------
    def generate_order_id(self):
        if not os.path.exists(self.order_file):
            return 1

        with open(self.order_file, "r") as file:
            return len(file.readlines()) + 1

    # ---------- VIEW RESERVED TABLES ----------
    def view_reserved_tables(self):
        self.reservation_obj.view_reservations()

    # ---------- TAKE ORDER ----------
    def take_order(self):
        print("\n------ TAKE ORDER ------")

        # Step 1: Show reserved tables
        self.view_reserved_tables()

        if not self.reservation_obj.reservations:
            print("❌ No reserved tables available.")
            return

        # Step 2: Select table
        table = input("Enter Table Number: ")

        reserved_tables = [str(i + 1) for i in range(len(self.reservation_obj.reservations))]

        if table not in reserved_tables:
            print("❌ Invalid table! Please select from reserved tables.")
            return

        # Step 3: Show menu
        print("\n📋 MENU:")
        self.menu_obj.display_menu()

        # Step 4: Take order input
        customer_name = input("Enter Customer Name: ")
        order_id = self.generate_order_id()

        items = []

        while True:
            item = input("Enter Item Name (or 'done'): ").strip()

            if item.lower() == "done":
                break

            try:
                qty = int(input("Enter Quantity: "))
            except ValueError:
                print("❌ Invalid quantity.")
                continue

            items.append((item, qty))

        if not items:
            print("⚠️ No items selected.")
            return

        # Step 5: Save order data
        self.save_order(order_id, customer_name, table, items)

        # Generate bill
        wb, ws = self.menu_obj.load_sheet()
        menu_dict = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            item_name, price, _, _ = row
            if item_name:
                menu_dict[item_name] = price

        order_dict = {}
        for item, qty in items:
            match = next((key for key in menu_dict if key.lower() == item.lower()), None)
            if match:
                order_dict[match] = order_dict.get(match, 0) + qty
            else:
                print(f"❌ {item} not found in menu, skipped.")

        if order_dict:
            generate_bill(order_dict, menu_dict)
        else:
            print("⚠️ No valid items to bill.")

        # ← NEW: Send review request email to customer
        selected_reservation = self.reservation_obj.reservations[int(table) - 1]
        customer_email = selected_reservation.get("email", "")
        if customer_email:
            review_request_email(customer_name, customer_email, order_id)

        # Remove the reserved table once the order is taken
        del self.reservation_obj.reservations[int(table) - 1]
        print(f"✅ Reservation for table {table} has been cleared.")

    # ---------- SAVE ORDER ----------
    def save_order(self, order_id, name, table, items):
        with open(self.order_file, "a", encoding="utf-8") as file:
            items_str = ";".join([f"{i}:{q}" for i, q in items])
            file.write(f"{order_id},{name},{table},{items_str}\n")

        self.save_order_xlsx(order_id, name, table, items)
        print("✅ Order saved successfully to TXT and XLSX.")

    def save_order_xlsx(self, order_id, name, table, items):
        try:
            wb = load_workbook(self.order_xlsx)
            ws = wb.active
            items_str = ";".join([f"{i}:{q}" for i, q in items])
            ws.append([order_id, name, table, items_str])
            wb.save(self.order_xlsx)
        except PermissionError:
            print("❌ Unable to update orders.xlsx because the file is open. Close it and try again.")

    # ---------- VIEW ORDERS ----------
    def view_orders(self):
        print("\n------ ORDER HISTORY ------")

        if not os.path.exists(self.order_file):
            print("No orders found.")
            return

        with open(self.order_file, "r") as file:
            for line in file:
                oid, name, table, items = line.strip().split(",")

                print(f"\nOrder ID: {oid}")
                print(f"Customer: {name}")
                print(f"Table: {table}")

                for i in items.split(";"):
                    item, qty = i.split(":")
                    print(f"  - {item} x {qty}")